import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import time 

workbook = load_workbook(filename="4BeatsQ1.xlsx")

current_day = datetime.datetime.now().strftime("%A")


sheet = workbook[current_day]

driver = webdriver.Firefox() #In case of using chorme replace FireFox() with Chrome()

def find_longest_shortest(keyword):
    driver.get("https://www.google.com")
    search_box = driver.find_element(By.NAME, 'q')
    search_box.send_keys(keyword)
    time.sleep(2)

    
    suggestions = driver.find_elements(By.CSS_SELECTOR, 'li.sbct span')
    options = [elem.text for elem in suggestions if elem.text]
    
    print(options)
    if options:
        longest_option = max(options)
        shortest_option = min(options)
    else:
        longest_option = ""
        shortest_option = ""
    return longest_option, shortest_option

for row_idx, row in enumerate(sheet.iter_rows(min_row=3, max_row=12, min_col=3, max_col=3), start=3): #hard coded SHOULD change later
    keyword = row[0].value  
    if keyword:
        longest, shortest = find_longest_shortest(keyword)
        print(len(longest),len(shortest))
        sheet.cell(row=row_idx, column=4).value = longest
        sheet.cell(row=row_idx, column=5).value = shortest


workbook.save("updated_excel_file.xlsx")

driver.quit()
